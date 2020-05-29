#从表格中取出IP
#1.新增IP
#2.减少IP

from openpyxl import load_workbook

def write(m):
    file = open('EternalBlueSMB-IP.csv', 'a+')
    file.write(m + '\n')
    file.close()

def main():
    ##########读出IP并处理
    txt1=[]
    txt2=[]
    txt3=[]
    count1 =0
    count2 = 0
    count3 =0
    for j in range(0,1):
        #addr = "D:\\Py-OP/2387.xlsx"
        addr2="D:\\Python-OP/Warnning/ip_list.xlsx"
        wb = load_workbook(addr2)
        sheet = wb.get_sheet_by_name("1")
        #print("sheet-1")
        for i in range(2, sheet.max_row):
            m = sheet["A" + str(i)].value
            m2 = m
            if m2 in txt1:
                #添加仍未处理的IP
                #xt3.append(m2)
                #print(m2)
                continue
            txt1.append(m2)
            #print(m2)

        #print("------------------sheet-2-----------------")
        sheet2 = wb.get_sheet_by_name("2")
        for j in range(2,sheet2.max_row):
            n = sheet2["A"+ str(j)].value
            txt2.append(n)
            #print(n)

    ##########新增IP
        for k in txt1:
            if k in txt2:
                #print(k + ":")
                txt3.append(k)
                continue
            print("IP:" + k + ":" + "该IP为新增IP")
            count1 +=1
        print("新增IP " + str(count1) + " 个，建议处理\n")
        print("--------------------------------------------------")
    #########减少IP
        for o in txt2:
            if o in txt1:
                continue
            print(o + "：该IP已被处理")
            count2 +=1
        print("本次处理后" + str(count2) + "个IP未复发。")
        print("--------------------------------------------------")

        print("顽固IP如下：")
        for q in txt3:
            print(q)
            count3 +=1
        print("顽固IP共：" + str(count3) + " 个。")
if __name__ == '__main__':
    main()
