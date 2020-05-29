from openpyxl import load_workbook
#import sys
def write(m):
    file = open('0416.csv', 'a+')
    file.write(m + '\n')
    file.close()

def main():
    ##########读出IP并处理
    txt=[]
    for j in range(0,1):
        #addr = "D:\\Py-OP/2387.xlsx"
        addr2="D:\\Python-OP/Warnning/20200416.xlsx"
        wb = load_workbook(addr2)
        sheet = wb.get_sheet_by_name("告警")

        for i in range(2, sheet.max_row):
            m = sheet["B" + str(i)].value
            m2 = m[2:-3]
            if m2 in txt:
                continue
            txt.append(m2)
            print(m2)

            write(m2)


    ##########写入IP


if __name__ == '__main__':
    #arg = sys.argv
    main()
