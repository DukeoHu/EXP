#########################从告警中提取出IP，并生成csv文件
####从XLSX中获取IP
from openpyxl import load_workbook

def write(m):
    file = open('EternalBlueSMB-IP.csv', 'a+')
    file.write(m + '\n')
    file.close()

def main():
    ##########读出IP并处理
    for j in range(0,1):
        addr = "D:\\Py-OP\Warnning/" + str(j) + "0000.xlsx"
        addr2="D:\\Py-OP\Warnning/2387.xls"
        wb = load_workbook(addr2)
        sheet = wb.get_sheet_by_name("告警")

        for i in range(2, sheet.max_row):
            m = sheet["P" + str(i)].value
            print(m[2:-2])
            m = m[2:-2]
            write(m)


    ##########写入IP


if __name__ == '__main__':
    main()
